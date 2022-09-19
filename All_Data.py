from openpyxl import load_workbook
import requests


class DataFromTXT:

    def __init__(self, txt_filename):
        self.txt_filename = txt_filename
        self.data_from_file = DataFromTXT.getDataFromFile(self)

    def getDataFromFile(self):
        with open(self.txt_filename, "r") as data:
            return data.read().splitlines()

    def createListFromDataFromFile(self, index):
        result_list = []
        for element in self.data_from_file:
            result_list.append(element.split()[index])
        return result_list

    def createDictFromDataDependsOnFile(self, keys_list, values_list):
        result_dict = {}
        i = 0
        for key in keys_list:
            result_dict[key] = values_list[i]
            i += 1
        return result_dict


class DataFromExcel:

    def __init__(self, excel_name):
        self.excel_name = excel_name


    def getValuesFromExcelToList(self, start_row, column_number):
        workbook = load_workbook(self.excel_name)
        workbook = workbook.active
        result_list = []
        while True:
            crypto = workbook.cell(row=start_row, column=column_number).value
            start_row += 1
            if crypto is None:
                break
            else:
                result_list.append(crypto)
        return result_list


class API_Data:

    def __init__(self, url='https://api.binance.com/api/v3/ticker/24hr'):
        self.url = url
        self.dataFromApi = API_Data.getDataFromAPI(self)
        self.all_tickets_from_Api = self.dataFromApi[0].keys()


    def getDataFromAPI(self): # use to check tickets names
        return requests.get(self.url).json()


    def createCryptoDictWithValues(self, crypto_dict, value_ticket, crypto_ticket):
        result_dict = {}
        for crypto in crypto_dict:
            for element in self.dataFromApi:
                if element[crypto_ticket] == crypto:
                    result_dict[crypto] = float(element[value_ticket])
        return result_dict


class CryptoData:

    def __init__(self, crypto_list, crypto_amount_list, crypto_price_transactions_list):
        self.crypto_list = crypto_list
        self.crypto_dict = dict.fromkeys(self.crypto_list)
        self.crypto_all_transactions_dict = CryptoData.createCryptoDictWithValues(self, crypto_price_transactions_list)
        self.crypto_amount_dict = CryptoData.createCryptoDictWithValues(self, crypto_amount_list)


    def createCryptoDictWithValues(self, crypto_values_list):
        result_dict = dict.fromkeys(self.crypto_list)
        for crypto in result_dict:
            result_dict[crypto] = sum([float(crypto_values_list[i]) for i in range(len(crypto_values_list)) if self.crypto_list[i] == crypto])
        return result_dict


class FinishData:

    def __init__(self,
                 output_file_name,
                 start_crypto_list, transactions_crypto_list,
                 start_total_cost_dict, transactions_total_cost_dict,
                 start_amount_dict, transactions_amount_list_dict,
                 ):

        self.output_file_name = output_file_name
        self.crypto_list = list(set(start_crypto_list + transactions_crypto_list))
        self.crypto_dict = dict.fromkeys(self.crypto_list)
        self.crypto_amount_dict = FinishData.sumStartDataAndNewData(self, start_amount_dict, transactions_amount_list_dict)
        self.crypto_total_cost_dict = FinishData.sumStartDataAndNewData(self, start_total_cost_dict, transactions_total_cost_dict)
        self.crypto_average_price_dict = FinishData.countCryptoAveragePrice(self) #, self.crypto_list, self.crypto_total_cost_dict, self.crypto_amount_dict)
        self.total_cost = sum(list(self.crypto_total_cost_dict.values()))
        self.crypto_percent_cost = FinishData.countPortfelPercentCost(self)#, user_dictionary=self.crypto_total_cost_dict)


    def sumStartDataAndNewData(self, start_data_dict, new_data_dict):
        for element in new_data_dict:
            if element in start_data_dict.keys():
                start_data_dict[element] = start_data_dict[element] + new_data_dict[element]
        return start_data_dict


    def countCryptoAveragePrice(self): #crypto_list, crypto_costs, crypto_amount):
        crypto_average_prices = {}
        for crypto in self.crypto_list:
            crypto_average_prices[crypto] = round((self.crypto_total_cost_dict[crypto] / self.crypto_amount_dict[crypto]),3)
        return crypto_average_prices


    def countCryptoProfit(self, actual_crypto_prices_dict):
        crypto_profit = {}
        for crypto in self.crypto_dict:
            crypto_profit[crypto] = (100 * actual_crypto_prices_dict[crypto] / self.crypto_average_price_dict[crypto]) - 100
        return crypto_profit


    def countPortfelPercentCost(self):#, user_dictionary):
        crypto_percent_cost = {}
        for crypto in self.crypto_total_cost_dict:
            crypto_percent_cost[crypto] = self.crypto_total_cost_dict[crypto] * 100 / self.total_cost
        return crypto_percent_cost


    def save_data_to_txt(self, actual_crypto_price_dict, crypto_profit_dict):
        with open(self.output_file_name, 'w') as report_file:
            report_file.write('{0:20} {1:20} {2:20} {3:20} {4:25} {5:25} {6:25}\n'.format('Crypto', 'Amount', 'Total Costs USDT', 'Portfel Percent', 'AveragePrice', 'Crypto Price', 'Profit%'))
            for crypto in self.crypto_dict:
                report_file.write('{0:5}'
                                  ' {1:20.4f}'
                                  ' {2:20.3f}'
                                  ' {3:20.3f}'
                                  ' {4:20.3f}'
                                  ' {5:20.3f}'
                                  ' {6:20.3f}\n'.format(crypto, float(self.crypto_amount_dict[crypto]),
                                                        float(self.crypto_total_cost_dict[crypto]),
                                                        self.crypto_percent_cost[crypto],
                                                        self.crypto_average_price_dict[crypto],
                                                        actual_crypto_price_dict[crypto],
                                                        crypto_profit_dict[crypto]))
            report_file.close()


class PotentialInvestment:


    def __init__(self,
                 actual_crypto_cost_dict,
                 crypto_coefficients_dict,
                 investment):

        self.crypto_coefficients_dict = crypto_coefficients_dict
        self.actual_crypto_cost_dict = actual_crypto_cost_dict
        self.crypto_costs_dict = PotentialInvestment.createCryptoCostsDict(self)
        self.actual_total_cost = sum(list(self.crypto_costs_dict.values()))
        self.investment = investment
        self.potential_total_investment = self.actual_total_cost + self.investment
        self.percent_structure = list(set(self.crypto_coefficients_dict.values()))
        self.actual_portfel_situation_main = PotentialInvestment.countPortfelSituationByCoefficient(self)
        self.expected_portfel_situation = PotentialInvestment.countExpectedPortfelSitation(self)
        self.investment_divided_by_percent = PotentialInvestment.createInvestmentDivivedByPercent(self)
        self.coefficient_for_crypto_dict = PotentialInvestment.createSummaryCryptoCoefficien(self)


    def countPortfelSituationByCoefficient(self):
        portfel_situation_by_coefficint_dict = {}
        for percent in self.percent_structure:
            portfel_situation_by_coefficint_dict[percent] = 0
            for crypto in self.crypto_coefficients_dict:
                if self.crypto_coefficients_dict[crypto] == percent:
                    portfel_situation_by_coefficint_dict[percent] += self.crypto_costs_dict[crypto]
                else:
                    continue
        return portfel_situation_by_coefficint_dict


    def countExpectedPortfelSitation(self):
        expected_portfel_sitation = {}
        for coefficient in self.percent_structure:
            expected_portfel_sitation[coefficient] = float(coefficient) * self.potential_total_investment
        return expected_portfel_sitation


    def createInvestmentDivivedByPercent(self):
        expected_portfel_sitation = {}
        for coefficient in self.percent_structure:
            expected_portfel_sitation[coefficient] = self.expected_portfel_situation[coefficient] - self.actual_portfel_situation_main[coefficient]
        return expected_portfel_sitation


    def createSummaryCryptoCoefficien(self):
        summary_crypto_coefficient = {}
        for coefficient in self.percent_structure:
            crypto_list = []
            for crypto in self.crypto_coefficients_dict:
                if self.crypto_coefficients_dict[crypto] == coefficient:
                    crypto_list.append(crypto)
            summary_crypto_coefficient[coefficient] = crypto_list
        return summary_crypto_coefficient


    def createCryptoCostsDict(self):
        crypto_cots_dict = {}
        for crypto in self.crypto_coefficients_dict:
            crypto_cots_dict[crypto] = self.actual_crypto_cost_dict[crypto]
        return crypto_cots_dict


    def createPotencialInvestmentByCrypto(self):
        crypto_cost_dict = {}
        for coefficient in self.percent_structure:
            crypto_cost_dict[coefficient] = self.investment_divided_by_percent[coefficient] / len(self.coefficient_for_crypto_dict[coefficient])
        return crypto_cost_dict


class CryptoCoefficient:

    def __init__(self,
                 crypto_coefficient_dict,
                 crypto_base_filename_2_lvl = ''):

        self.crypto_coefficient_dict = crypto_coefficient_dict
        self.crypto_base_filename_2_lvl = crypto_base_filename_2_lvl
        self.crypto_list = list(self.crypto_coefficient_dict.keys())
        self.structure = set(self.crypto_coefficient_dict.values())


    def createCrypto_coefficient_file(self, crypto_list_file_name, crypto_list):
        with open(crypto_list_file_name, "w") as my_coefficient_file:
            for crypto in crypto_list:
                my_coefficient_file.write(crypto + '\n')

    def createCrypto_coefficient_files_by_coefficient(self):
        for coefficient in self.structure:
            with open((str(coefficient) + '_' + self.crypto_base_filename_2_lvl), "w") as my_coefficient_file:
                for crypto in self.crypto_coefficient_dict:
                    if self.crypto_coefficient_dict[crypto] == coefficient:
                        my_coefficient_file.write(crypto + '\n')


